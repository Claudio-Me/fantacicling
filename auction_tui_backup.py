#!/usr/bin/env python3
"""Fantasy Cycling Auction TUI - Modern terminal UI with Textual."""

import csv
import sys
import os
from openpyxl import load_workbook
from textual.app import App, ComposeResult
from textual.binding import Binding
from textual.containers import Container, Horizontal, Vertical, Center
from textual.reactive import reactive
from textual.screen import ModalScreen
from textual.widgets import Button, Footer, Header, Input, Label, ProgressBar, Static


# ============================================================================
# Data Functions (reused from auction.py)
# ============================================================================

def read_riders_from_excel(filepath):
    """Read rider names and values from Excel file (surname in col B, first name in col C, value in col G)."""
    wb = load_workbook(filepath, read_only=True)
    ws = wb.active
    riders = []
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=7, values_only=True):
        surname = str(row[0]).strip() if row[0] else ''
        firstname = str(row[1]).strip() if row[1] else ''
        value = row[5] if len(row) > 5 and row[5] is not None else None  # col G (index 5 from col B)
        if surname or firstname:
            rider_name = f"{firstname} {surname}".strip()
            if rider_name:
                riders.append({'name': rider_name, 'value': value})
    wb.close()
    return riders


def read_riders_from_csv(filepath):
    """Read rider names from CSV file (first column). Value is None for CSV files."""
    riders = []
    with open(filepath, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        for row in reader:
            if row and row[0].strip():
                riders.append({'name': row[0].strip(), 'value': None})
    return riders


def read_riders(filepath):
    """Read riders from file (Excel or CSV)."""
    if filepath.endswith('.xlsx') or filepath.endswith('.xls'):
        return read_riders_from_excel(filepath)
    elif filepath.endswith('.csv'):
        return read_riders_from_csv(filepath)
    else:
        raise ValueError(f"Formato file non supportato: {filepath}")


def save_results(results, riders, output_file):
    """Save auction results to CSV."""
    with open(output_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(['Corridore', 'Squadra', 'Prezzo'])
        for i, rider in enumerate(riders):
            rider_name = rider['name'] if isinstance(rider, dict) else rider
            team = results[i].get('team', '')
            price = results[i].get('price', '')
            writer.writerow([rider_name, team if team else '', price if price else ''])


# ============================================================================
# CSS Styling
# ============================================================================

CSS = """
Screen {
    background: $surface;
}

#main-container {
    width: 100%;
    height: 100%;
    padding: 1 2;
}

#rider-panel {
    width: 100%;
    height: auto;
    border: solid $primary;
    padding: 2 3;
    margin: 2 0;
    background: $surface-lighten-1;
}

#rider-number {
    text-style: bold;
    color: $text-muted;
    margin-bottom: 1;
}

#rider-name {
    text-style: bold;
    color: $primary-lighten-2;
    text-align: center;
    width: 100%;
    margin: 1 0;
}

#rider-value {
    text-align: center;
    width: 100%;
    color: $text-muted;
    margin-top: 1;
}

#progress-container {
    width: 100%;
    height: auto;
    margin: 1 0;
}

#progress-bar {
    width: 100%;
    padding: 0 1;
}

#progress-label {
    text-align: right;
    width: 100%;
    color: $text-muted;
}

#status-panel {
    width: 100%;
    height: auto;
    padding: 1 2;
    margin: 1 0;
    border: solid $secondary;
    background: $surface-lighten-1;
}

#status-label {
    text-align: center;
    width: 100%;
}

.status-assigned {
    color: $success;
}

.status-unassigned {
    color: $warning;
}

#button-bar {
    width: 100%;
    height: auto;
    margin: 2 0 1 0;
    align: center middle;
}

#button-bar Button {
    margin: 0 1;
}

Button.primary {
    background: $primary;
}

Button.success {
    background: $success;
}

Button.warning {
    background: $warning;
}

/* Modal Styles */
AssignmentModal {
    align: center middle;
}

#modal-container {
    width: 60;
    height: auto;
    border: thick $primary;
    background: $surface;
    padding: 1 2;
}

#modal-title {
    text-style: bold;
    text-align: center;
    width: 100%;
    margin-bottom: 1;
    color: $primary-lighten-2;
}

#modal-rider-name {
    text-align: center;
    width: 100%;
    margin-bottom: 1;
    color: $text;
    text-style: bold;
}

.input-label {
    margin-top: 1;
    margin-bottom: 0;
}

.modal-input {
    width: 100%;
    margin-bottom: 1;
}

#modal-buttons {
    margin-top: 1;
    align: center middle;
}

#modal-buttons Button {
    margin: 0 1;
}
"""


# ============================================================================
# Assignment Modal
# ============================================================================

class AssignmentModal(ModalScreen[tuple[str, str] | None]):
    """Modal dialog for team/price input."""

    BINDINGS = [
        Binding("escape", "cancel", "Annulla"),
    ]

    def __init__(self, rider_name: str) -> None:
        super().__init__()
        self.rider_name = rider_name

    def compose(self) -> ComposeResult:
        with Container(id="modal-container"):
            yield Label("Assegna Corridore", id="modal-title")
            yield Label(self.rider_name, id="modal-rider-name")
            yield Label("Squadra:", classes="input-label")
            yield Input(placeholder="Nome squadra (vuoto per saltare)", id="team-input", classes="modal-input")
            yield Label("Prezzo:", classes="input-label")
            yield Input(placeholder="Prezzo (es. 100)", id="price-input", classes="modal-input")
            with Horizontal(id="modal-buttons"):
                yield Button("Conferma", variant="success", id="confirm-btn")
                yield Button("Annulla", variant="error", id="cancel-btn")

    def on_mount(self) -> None:
        self.query_one("#team-input", Input).focus()

    def on_button_pressed(self, event: Button.Pressed) -> None:
        if event.button.id == "confirm-btn":
            self._confirm()
        elif event.button.id == "cancel-btn":
            self.dismiss(None)

    def on_input_submitted(self, event: Input.Submitted) -> None:
        if event.input.id == "team-input":
            self.query_one("#price-input", Input).focus()
        elif event.input.id == "price-input":
            self._confirm()

    def _confirm(self) -> None:
        team = self.query_one("#team-input", Input).value.strip()
        price = self.query_one("#price-input", Input).value.strip()
        if team:
            self.dismiss((team, price))
        else:
            self.dismiss(None)

    def action_cancel(self) -> None:
        self.dismiss(None)


# ============================================================================
# Main Application
# ============================================================================

class AuctionApp(App):
    """Main auction TUI application."""

    CSS = CSS
    TITLE = "ASTA FANTACICLISMO"

    BINDINGS = [
        Binding("up", "previous_rider", "Precedente", key_display="↑"),
        Binding("k", "previous_rider", "Precedente", show=False),
        Binding("down", "next_rider", "Successivo", key_display="↓"),
        Binding("j", "next_rider", "Successivo", show=False),
        Binding("enter", "assign", "Assegna", key_display="⏎"),
        Binding("q", "save_quit", "Salva ed Esci"),
    ]

    current_index: reactive[int] = reactive(0)

    def __init__(self, riders: list[str], output_file: str) -> None:
        super().__init__()
        self.riders = riders
        self.output_file = output_file
        self.total = len(riders)
        self.results = [{'team': None, 'price': None} for _ in riders]

    def compose(self) -> ComposeResult:
        yield Header()
        with Container(id="main-container"):
            with Container(id="rider-panel"):
                yield Label("Corridore 1/1", id="rider-number")
                yield Label("Nome Corridore", id="rider-name")
                yield Label("", id="rider-value")
            with Container(id="progress-container"):
                yield ProgressBar(total=self.total, show_eta=False, id="progress-bar")
                yield Label("1/1", id="progress-label")
            with Container(id="status-panel"):
                yield Label("Stato: Non ancora assegnato", id="status-label", classes="status-unassigned")
            with Horizontal(id="button-bar"):
                yield Button("Precedente", id="prev-btn", variant="default")
                yield Button("Successivo", id="next-btn", variant="primary")
                yield Button("Assegna", id="assign-btn", variant="success")
                yield Button("Salva ed Esci", id="quit-btn", variant="warning")
        yield Footer()

    def on_mount(self) -> None:
        self._update_display()

    def watch_current_index(self, value: int) -> None:
        self._update_display()

    def _update_display(self) -> None:
        """Update all display elements based on current index."""
        idx = self.current_index
        rider = self.riders[idx]
        rider_name_str = rider['name'] if isinstance(rider, dict) else rider
        rider_value = rider.get('value') if isinstance(rider, dict) else None
        result = self.results[idx]

        # Update rider number
        rider_number = self.query_one("#rider-number", Label)
        rider_number.update(f"Corridore {idx + 1}/{self.total}")

        # Update rider name
        rider_name = self.query_one("#rider-name", Label)
        rider_name.update(rider_name_str)

        # Update rider value
        value_label = self.query_one("#rider-value", Label)
        if rider_value is not None:
            value_label.update(f"Valore: {rider_value}")
        else:
            value_label.update("")

        # Update progress bar
        progress_bar = self.query_one("#progress-bar", ProgressBar)
        progress_bar.progress = idx + 1

        # Update progress label
        progress_label = self.query_one("#progress-label", Label)
        progress_label.update(f"{idx + 1}/{self.total}")

        # Update status
        status_label = self.query_one("#status-label", Label)
        if result['team']:
            status_label.update(f"Assegnato: {result['team']} - {result['price']}")
            status_label.remove_class("status-unassigned")
            status_label.add_class("status-assigned")
        else:
            status_label.update("Stato: Non ancora assegnato")
            status_label.remove_class("status-assigned")
            status_label.add_class("status-unassigned")

    def on_button_pressed(self, event: Button.Pressed) -> None:
        """Handle button clicks."""
        if event.button.id == "prev-btn":
            self.action_previous_rider()
        elif event.button.id == "next-btn":
            self.action_next_rider()
        elif event.button.id == "assign-btn":
            self.action_assign()
        elif event.button.id == "quit-btn":
            self.action_save_quit()

    def action_previous_rider(self) -> None:
        """Go to previous rider."""
        if self.current_index > 0:
            self.current_index -= 1

    def action_next_rider(self) -> None:
        """Go to next rider."""
        if self.current_index < self.total - 1:
            self.current_index += 1

    def action_assign(self) -> None:
        """Open assignment modal."""
        rider = self.riders[self.current_index]
        rider_name = rider['name'] if isinstance(rider, dict) else rider
        self.push_screen(AssignmentModal(rider_name), self._handle_assignment)

    def _handle_assignment(self, result: tuple[str, str] | None) -> None:
        """Handle modal result."""
        if result:
            team, price = result
            self.results[self.current_index]['team'] = team
            self.results[self.current_index]['price'] = price
            self._update_display()
            # Auto-advance to next rider
            if self.current_index < self.total - 1:
                self.current_index += 1

    def action_save_quit(self) -> None:
        """Save results and quit."""
        save_results(self.results, self.riders, self.output_file)
        assigned = sum(1 for r in self.results if r['team'])
        self.exit(message=f"Risultati salvati in: {self.output_file}\nRiepilogo: {assigned}/{self.total} corridori assegnati")


# ============================================================================
# Main Entry Point
# ============================================================================

def get_unique_output_file(base_name):
    """Generate unique filename, adding _1, _2, etc. if file exists."""
    output_file = f"{base_name}_auction_results.csv"
    if not os.path.exists(output_file):
        return output_file

    counter = 1
    while True:
        output_file = f"{base_name}_auction_results_{counter}.csv"
        if not os.path.exists(output_file):
            return output_file
        counter += 1


def main():
    if len(sys.argv) < 2:
        print("Utilizzo: python auction_tui.py <file_input.xlsx>")
        print("         python auction_tui.py <file_input.csv>")
        sys.exit(1)

    input_file = sys.argv[1]
    if not os.path.exists(input_file):
        print(f"Errore: File non trovato: {input_file}")
        sys.exit(1)

    # Determine output file name (unique to avoid overwriting)
    base_name = os.path.splitext(os.path.basename(input_file))[0]
    output_file = get_unique_output_file(base_name)

    # Read riders
    print(f"Lettura corridori da: {input_file}")
    riders = read_riders(input_file)
    total = len(riders)
    print(f"Trovati {total} corridori")

    if total == 0:
        print("Nessun corridore trovato nel file!")
        sys.exit(1)

    # Run the TUI app
    app = AuctionApp(riders, output_file)
    result = app.run()
    if result:
        print(result)


if __name__ == "__main__":
    main()
