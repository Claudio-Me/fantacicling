#!/usr/bin/env python3
"""Fantasy Cycling Auction Script - Assign riders to teams with prices."""

import csv
import sys
import os
from openpyxl import load_workbook
import readchar
from rich.console import Console
from rich.panel import Panel
from rich.table import Table
from rich.progress import Progress, BarColumn, TextColumn
from rich.text import Text
from rich import box

console = Console()


def read_riders_from_excel(filepath):
    """Read rider names from Excel file (surname in col B, first name in col C)."""
    wb = load_workbook(filepath, read_only=True)
    ws = wb.active
    riders = []
    # Skip header row (min_row=2), read columns B (2) and C (3)
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=3, values_only=True):
        surname = str(row[0]).strip() if row[0] else ''
        firstname = str(row[1]).strip() if row[1] else ''
        if surname or firstname:
            # Format: "Firstname SURNAME" (e.g., "Tadej POGACAR")
            rider = f"{firstname} {surname}".strip()
            if rider:
                riders.append(rider)
    wb.close()
    return riders


def read_riders_from_csv(filepath):
    """Read rider names from CSV file (first column)."""
    riders = []
    with open(filepath, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        for row in reader:
            if row and row[0].strip():
                riders.append(row[0].strip())
    return riders


def read_riders(filepath):
    """Read riders from file (Excel or CSV)."""
    if filepath.endswith('.xlsx') or filepath.endswith('.xls'):
        return read_riders_from_excel(filepath)
    elif filepath.endswith('.csv'):
        return read_riders_from_csv(filepath)
    else:
        raise ValueError(f"Formato file non supportato: {filepath}")


def create_progress_bar(current, total, width=40):
    """Create a text-based progress bar."""
    filled = int(width * current / total)
    empty = width - filled
    bar = "[green]" + "█" * filled + "[/green][dim]" + "░" * empty + "[/dim]"
    return bar


def display_rider(index, total, rider, current_team, current_price):
    """Display current rider info with Rich formatting."""
    console.clear()

    # Build main content
    progress_bar = create_progress_bar(index + 1, total)

    # Status text
    if current_team:
        status_text = f"[green]Assegnazione attuale: {current_team} - {current_price}[/green]"
    else:
        status_text = "[yellow]Non ancora assegnato[/yellow]"

    # Build the main panel content
    content = Text()
    content.append(f"Corridore {index + 1}/{total}: ", style="bold")
    content.append(f"{rider}\n\n", style="bold cyan")

    main_text = f"""[bold]Corridore {index + 1}/{total}:[/bold] [bold cyan]{rider}[/bold cyan]

{progress_bar} {index + 1}/{total}

{status_text}"""

    # Create main panel
    main_panel = Panel(
        main_text,
        title="[bold yellow]🚴 ASTA FANTACICLISMO 🚴[/bold yellow]",
        border_style="yellow",
        box=box.ROUNDED,
        padding=(1, 2)
    )
    console.print(main_panel)

    # Create commands table
    commands_table = Table(
        show_header=False,
        box=box.SIMPLE,
        padding=(0, 2),
        show_edge=False
    )
    commands_table.add_column("Key", style="bold magenta", width=8)
    commands_table.add_column("Description", style="white")

    commands_table.add_row("↑", "Vai al corridore precedente")
    commands_table.add_row("↓", "Vai al corridore successivo")
    commands_table.add_row("⏎", "Assegna squadra e prezzo")
    commands_table.add_row("q", "Salva ed esci")

    commands_panel = Panel(
        commands_table,
        title="[bold]Comandi[/bold]",
        border_style="blue",
        box=box.ROUNDED
    )
    console.print(commands_panel)


def get_team_and_price():
    """Prompt user for team name and price."""
    console.print("\n[bold cyan]Inserisci nome squadra[/bold cyan] (oppure 'salta' / premi Invio per saltare):")
    team = input("> ").strip()

    if not team or team.lower() in ('salta', 'skip'):
        return None, None

    console.print("[bold cyan]Inserisci prezzo:[/bold cyan]")
    price_str = input("> ").strip()

    try:
        price = int(price_str) if price_str else 0
    except ValueError:
        price = price_str  # Keep as string if not a number

    return team, price


def save_results(results, riders, output_file):
    """Save auction results to CSV."""
    with open(output_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(['Corridore', 'Squadra', 'Prezzo'])
        for i, rider in enumerate(riders):
            team = results[i].get('team', '')
            price = results[i].get('price', '')
            writer.writerow([rider, team if team else '', price if price else ''])
    console.print(f"\n[bold green]Risultati salvati in:[/bold green] {output_file}")


def main():
    if len(sys.argv) < 2:
        console.print("[bold red]Utilizzo:[/bold red] python auction.py <file_input.xlsx>")
        console.print("         python auction.py <file_input.csv>")
        sys.exit(1)

    input_file = sys.argv[1]
    if not os.path.exists(input_file):
        console.print(f"[bold red]Errore:[/bold red] File non trovato: {input_file}")
        sys.exit(1)

    # Determine output file name
    base_name = os.path.splitext(os.path.basename(input_file))[0]
    output_file = f"{base_name}_auction_results.csv"

    # Read riders
    console.print(f"[dim]Lettura corridori da:[/dim] {input_file}")
    riders = read_riders(input_file)
    total = len(riders)
    console.print(f"[dim]Trovati[/dim] [bold]{total}[/bold] [dim]corridori[/dim]")

    if total == 0:
        console.print("[bold red]Nessun corridore trovato nel file![/bold red]")
        sys.exit(1)

    # Initialize results
    results = [{'team': None, 'price': None} for _ in riders]

    # Main loop
    current_index = 0

    while True:
        rider = riders[current_index]
        current_team = results[current_index]['team']
        current_price = results[current_index]['price']

        display_rider(current_index, total, rider, current_team, current_price)

        # Wait for key press
        console.print("\n[dim]Premi un tasto...[/dim]")
        key = readchar.readkey()

        if key == readchar.key.UP:
            # Go to previous rider
            if current_index > 0:
                current_index -= 1
        elif key == readchar.key.DOWN:
            # Go to next rider
            if current_index < total - 1:
                current_index += 1
        elif key == '\r' or key == '\n':
            # Enter: assign team and price
            team, price = get_team_and_price()
            results[current_index]['team'] = team
            results[current_index]['price'] = price

            # Auto-advance to next rider
            if current_index < total - 1:
                current_index += 1
            else:
                # At the end, ask if done
                console.print("\n[yellow]Raggiunto l'ultimo corridore.[/yellow] Premi 'q' per salvare ed uscire, o un altro tasto per continuare.")
                key = readchar.readkey()
                if key == 'q':
                    break
        elif key == 'q':
            # Quit and save
            break

    # Save results
    save_results(results, riders, output_file)

    # Summary
    assigned = sum(1 for r in results if r['team'])
    console.print(f"\n[bold]Riepilogo:[/bold] [green]{assigned}/{total}[/green] corridori assegnati alle squadre")


if __name__ == "__main__":
    main()
